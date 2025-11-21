#!/usr/bin/env python3
"""
Extract all pricing data from the Excel file and output to JSON
"""
import openpyxl
import json
from decimal import Decimal

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'
OUTPUT_FILE = 'api/src/data/pricing-tables-complete.json'

def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError

def extract_materials(wb):
    """Extract all materials from the Materials sheet"""
    print("\n=== EXTRACTING MATERIALS ===")
    sheet = wb['Material']
    materials = []
    
    # Start from row 4 (headers in row 3)
    for row_idx in range(4, sheet.max_row + 1):
        code = sheet.cell(row=row_idx, column=2).value  # Column B
        name = sheet.cell(row=row_idx, column=3).value  # Column C
        unit = sheet.cell(row=row_idx, column=4).value  # Column D
        price = sheet.cell(row=row_idx, column=5).value  # Column E
        
        # Stop if we hit empty rows
        if not code and not name:
            break
            
        if code:  # Only add if we have a code
            materials.append({
                "code": str(code).strip() if code else "",
                "name": str(name).strip() if name else "",
                "unit": str(unit).strip() if unit else "",
                "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
            })
            print(f"Material: {code} - {name}")
    
    print(f"Total materials extracted: {len(materials)}")
    return materials

def extract_fittings(wb):
    """Extract all fittings from the Fittings sheet"""
    print("\n=== EXTRACTING FITTINGS ===")
    sheet = wb['Fittings']
    fittings = []
    
    # Start from row 4 (headers in row 3)
    for row_idx in range(4, sheet.max_row + 1):
        code = sheet.cell(row=row_idx, column=2).value  # Column B
        name = sheet.cell(row=row_idx, column=3).value  # Column C
        unit = sheet.cell(row=row_idx, column=4).value  # Column D
        price = sheet.cell(row=row_idx, column=5).value  # Column E
        
        # Stop if we hit empty rows
        if not code and not name:
            break
            
        if code:  # Only add if we have a code
            fittings.append({
                "code": str(code).strip() if code else "",
                "name": str(name).strip() if name else "",
                "unit": str(unit).strip() if unit else "",
                "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
            })
            print(f"Fitting: {code} - {name}")
    
    print(f"Total fittings extracted: {len(fittings)}")
    return fittings

def extract_edge_banding(wb):
    """Extract all edge banding from the Edge Banding sheet"""
    print("\n=== EXTRACTING EDGE BANDING ===")
    sheet = wb['NavarShiarFarsi']
    edge_banding = []
    
    # Start from row 4 (headers in row 3)
    for row_idx in range(4, sheet.max_row + 1):
        code = sheet.cell(row=row_idx, column=2).value  # Column B
        name = sheet.cell(row=row_idx, column=3).value  # Column C
        unit = sheet.cell(row=row_idx, column=4).value  # Column D
        price = sheet.cell(row=row_idx, column=5).value  # Column E
        
        # Stop if we hit empty rows
        if not code and not name:
            break
            
        if code:  # Only add if we have a code
            edge_banding.append({
                "code": str(code).strip() if code else "",
                "name": str(name).strip() if name else "",
                "unit": str(unit).strip() if unit else "",
                "pricePerMeter": float(price) if price and isinstance(price, (int, float)) else 0
            })
            print(f"Edge Banding: {code} - {name}")
    
    print(f"Total edge banding extracted: {len(edge_banding)}")
    return edge_banding

def extract_cnc_operations(wb):
    """Extract all CNC operations from the CNC sheet"""
    print("\n=== EXTRACTING CNC OPERATIONS ===")
    sheet = wb['CNC']
    cnc_operations = []
    
    # Start from row 4 (headers in row 3)
    for row_idx in range(4, sheet.max_row + 1):
        code = sheet.cell(row=row_idx, column=2).value  # Column B
        name = sheet.cell(row=row_idx, column=3).value  # Column C
        unit = sheet.cell(row=row_idx, column=4).value  # Column D
        price = sheet.cell(row=row_idx, column=5).value  # Column E
        
        # Stop if we hit empty rows
        if not code and not name:
            break
            
        if code:  # Only add if we have a code
            cnc_operations.append({
                "code": str(code).strip() if code else "",
                "name": str(name).strip() if name else "",
                "unit": str(unit).strip() if unit else "",
                "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
            })
            print(f"CNC Operation: {code} - {name}")
    
    print(f"Total CNC operations extracted: {len(cnc_operations)}")
    return cnc_operations

def extract_pricing_config(wb):
    """Extract pricing configuration from the Data sheet"""
    print("\n=== EXTRACTING PRICING CONFIG ===")
    sheet = wb['Data']
    
    # Based on previous analysis, extract key pricing factors
    config = {
        "laborCostPerHour": 0,
        "overheadPercentage": 0,
        "profitMarginPercentage": 0,
        "wastagePercentage": 5,  # Default
        "cncSetupCost": 0,
        "edgeBandingSetupCost": 0
    }
    
    # Try to find specific cells (you'll need to adjust based on actual Excel structure)
    # For now, return defaults
    print("Using default pricing config (needs manual adjustment)")
    return config

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    print(f"\nAvailable sheets: {wb.sheetnames}")
    
    data = {
        "materials": extract_materials(wb),
        "fittings": extract_fittings(wb),
        "edgeBanding": extract_edge_banding(wb),
        "cncOperations": extract_cnc_operations(wb),
        "pricingConfig": extract_pricing_config(wb)
    }
    
    # Write to JSON
    print(f"\n=== WRITING TO {OUTPUT_FILE} ===")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=decimal_default)
    
    print(f"\nExtraction complete!")
    print(f"Summary:")
    print(f"  - Materials: {len(data['materials'])}")
    print(f"  - Fittings: {len(data['fittings'])}")
    print(f"  - Edge Banding: {len(data['edgeBanding'])}")
    print(f"  - CNC Operations: {len(data['cncOperations'])}")

if __name__ == "__main__":
    main()

