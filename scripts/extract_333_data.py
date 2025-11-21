"""
Extract Data from 333.csv.xlsx Excel File
This script reads the Excel file and extracts:
1. Sheet structure (columns, data types)
2. Actual data rows
3. Column mapping documentation
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
from typing import Dict, List, Any

def extract_excel_data(filename: str) -> Dict[str, Any]:
    """
    Extract complete structure and data from Excel file
    """
    print(f"Loading {filename}...")
    wb = load_workbook(filename, data_only=True)
    
    result = {
        'sheet_names': wb.sheetnames,
        'sheets': {},
        'column_mapping': {},
    }
    
    # Process first sheet (should be the data sheet)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    print(f"Processing sheet: {sheet_name}")
    
    # Get max row and column
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"Dimensions: {max_row} rows x {max_col} columns")
    
    # Extract all data
    data = []
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            # First row - headers
            headers = [cell if cell else f"Column_{idx}" for idx, cell in enumerate(row, start=1)]
            print(f"Headers: {headers}")
        else:
            # Data rows
            row_data = list(row)
            # Only include non-empty rows
            if any(cell is not None and str(cell).strip() != '' for cell in row_data):
                data.append(row_data)
    
    result['sheets'][sheet_name] = {
        'headers': headers,
        'data': data,
        'row_count': len(data),
        'column_count': len(headers),
    }
    
    # Analyze data patterns
    result['analysis'] = analyze_data_patterns(headers, data)
    
    return result

def analyze_data_patterns(headers: List[str], data: List[List[Any]]) -> Dict[str, Any]:
    """
    Analyze the data to identify patterns
    """
    analysis = {
        'sample_rows': data[:5] if len(data) >= 5 else data,
        'column_types': {},
        'unique_values': {},
    }
    
    # Analyze each column
    for col_idx, header in enumerate(headers):
        column_values = [row[col_idx] if col_idx < len(row) else None for row in data]
        non_null_values = [v for v in column_values if v is not None]
        
        if non_null_values:
            # Determine column type
            first_val = non_null_values[0]
            col_type = type(first_val).__name__
            
            analysis['column_types'][header] = col_type
            
            # Get unique values for categorical columns
            if col_type == 'str' and len(set(non_null_values)) < 50:
                analysis['unique_values'][header] = list(set(non_null_values))[:20]
    
    return analysis

def save_as_csv(data_dict: Dict[str, Any], output_filename: str):
    """
    Save extracted data as clean CSV
    """
    sheet_name = list(data_dict['sheets'].keys())[0]
    sheet_data = data_dict['sheets'][sheet_name]
    
    with open(output_filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Write headers
        writer.writerow(sheet_data['headers'])
        
        # Write data
        for row in sheet_data['data']:
            writer.writerow(row)
    
    print(f"Saved clean CSV to: {output_filename}")

def save_column_mapping(data_dict: Dict[str, Any], output_filename: str):
    """
    Save column mapping documentation
    """
    sheet_name = list(data_dict['sheets'].keys())[0]
    sheet_data = data_dict['sheets'][sheet_name]
    analysis = data_dict['analysis']
    
    mapping = []
    mapping.append("# Column Mapping for 333.csv.xlsx\n")
    mapping.append(f"## Sheet: {sheet_name}\n")
    mapping.append(f"## Total Rows: {sheet_data['row_count']}\n")
    mapping.append(f"## Total Columns: {sheet_data['column_count']}\n\n")
    
    mapping.append("### Column Structure (Excel-style):\n\n")
    mapping.append("| Excel Col | Index | Header | Type | Description |\n")
    mapping.append("|-----------|-------|--------|------|-------------|\n")
    
    # Expected mapping based on chatgptreport
    expected_mapping = {
        0: ("A", "Name/Component", "Component name or description"),
        1: ("B", "Component ID", "Unique component identifier"),
        2: ("C", "Quantity", "Number of components"),
        3: ("D", "Edge1", "Edge property 1"),
        4: ("E", "Edge2", "Edge property 2"),
        5: ("F", "Edge3", "Edge property 3"),
        6: ("G", "Edge4", "Edge property 4"),
        7: ("H", "Material Type", "Material specification"),
        8: ("I", "Instance Type", "Instance name (may have CNC- prefix)"),
        9: ("J", "Length", "Length in mm"),
        10: ("K", "Width", "Width in mm"),
        11: ("L", "Area", "Area in m²"),
    }
    
    for idx, header in enumerate(sheet_data['headers']):
        excel_col = chr(65 + idx) if idx < 26 else f"{chr(65 + idx // 26 - 1)}{chr(65 + idx % 26)}"
        col_type = analysis['column_types'].get(header, 'unknown')
        
        if idx in expected_mapping:
            expected_col, expected_name, expected_desc = expected_mapping[idx]
            mapping.append(f"| {excel_col} | {idx} | {header} | {col_type} | {expected_desc} |\n")
        else:
            mapping.append(f"| {excel_col} | {idx} | {header} | {col_type} | - |\n")
    
    mapping.append("\n### Sample Data (First 5 Rows):\n\n")
    mapping.append("```\n")
    for row_idx, row in enumerate(analysis['sample_rows'], start=1):
        mapping.append(f"Row {row_idx}: {row}\n")
    mapping.append("```\n")
    
    mapping.append("\n### Unique Values (Categorical Columns):\n\n")
    for col, values in analysis['unique_values'].items():
        mapping.append(f"**{col}**: {', '.join(str(v) for v in values[:10])}\n")
    
    with open(output_filename, 'w', encoding='utf-8') as f:
        f.writelines(mapping)
    
    print(f"Saved column mapping to: {output_filename}")

def main():
    input_file = '333.csv.xlsx'
    
    try:
        # Extract data
        extracted_data = extract_excel_data(input_file)
        
        # Save as clean CSV
        save_as_csv(extracted_data, '333_clean.csv')
        
        # Save column mapping documentation
        save_column_mapping(extracted_data, 'COLUMN_MAPPING.md')
        
        # Save full analysis as JSON
        # Remove large data arrays for JSON (keep only structure)
        json_data = {
            'sheet_names': extracted_data['sheet_names'],
            'structure': {
                k: {
                    'headers': v['headers'],
                    'row_count': v['row_count'],
                    'column_count': v['column_count'],
                }
                for k, v in extracted_data['sheets'].items()
            },
            'analysis': {
                'column_types': extracted_data['analysis']['column_types'],
                'unique_values': extracted_data['analysis']['unique_values'],
                'sample_rows': extracted_data['analysis']['sample_rows'],
            }
        }
        
        with open('333_structure.json', 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print("\n✅ Extraction complete!")
        print("Generated files:")
        print("  - 333_clean.csv (Clean CSV data)")
        print("  - COLUMN_MAPPING.md (Column documentation)")
        print("  - 333_structure.json (Structure analysis)")
        
    except FileNotFoundError:
        print(f"❌ Error: File '{input_file}' not found!")
        print("Please ensure 333.csv.xlsx is in the current directory.")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

