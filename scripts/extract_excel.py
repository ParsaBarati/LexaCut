import os
import openpyxl
from oletools.olevba import VBA_Parser

# --- CONFIGURATION ---
FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'
OUTPUT_FILE = 'extraction_report.txt'

def extract_macros(file_path, output_handle):
    """Extracts VBA code from the binary vbaProject.bin structure."""
    output_handle.write(f"\n{'='*40}\nSECTION 1: VBA MACROS\n{'='*40}\n")
    
    vbaparser = VBA_Parser(file_path)
    
    if vbaparser.detect_vba_macros():
        output_handle.write("Status: VBA Macros detected. Extracting...\n\n")
        for (filename, stream_path, vba_filename, vba_code) in vbaparser.extract_macros():
            output_handle.write(f"--- MODULE: {vba_filename} ---\n")
            output_handle.write(f"{'-'*30}\n")
            output_handle.write(vba_code)
            output_handle.write(f"\n{'-'*30}\n\n")
    else:
        output_handle.write("Status: No VBA Macros found in this file.\n")
    
    vbaparser.close()

def extract_formulas(file_path, output_handle):
    """Extracts cell formulas using openpyxl."""
    output_handle.write(f"\n{'='*40}\nSECTION 2: EXCEL FORMULAS\n{'='*40}\n")
    
    try:
        # data_only=False ensures we get the FORMULA, not the calculated value
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output_handle.write(f"\n--- SHEET: {sheet_name} ---\n")
            found_formula = False
            
            for row in ws.iter_rows():
                for cell in row:
                    # Check if the cell contains a formula (starts with =)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        output_handle.write(f"Cell {cell.coordinate}: {cell.value}\n")
                        found_formula = True
            
            if not found_formula:
                output_handle.write("(No formulas found in this sheet)\n")
                
    except Exception as e:
        output_handle.write(f"Error extracting formulas: {str(e)}\n")

def main():
    if not os.path.exists(FILE_NAME):
        print(f"ERROR: The file '{FILE_NAME}' was not found in the current directory.")
        return

    print(f"Processing {FILE_NAME}...")
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(f"EXTRACTION REPORT FOR: {FILE_NAME}\n")
        
        # 1. Extract Macros
        try:
            extract_macros(FILE_NAME, f)
            print("Macros extracted.")
        except Exception as e:
            f.write(f"\nError processing macros: {e}\n")
            print(f"Error processing macros: {e}")

        # 2. Extract Formulas
        try:
            extract_formulas(FILE_NAME, f)
            print("Formulas extracted.")
        except Exception as e:
            f.write(f"\nError processing formulas: {e}\n")
            print(f"Error processing formulas: {e}")

    print(f"Done! Results saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
