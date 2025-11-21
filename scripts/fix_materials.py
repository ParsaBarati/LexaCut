#!/usr/bin/env python3
"""
Extract CORRECT material pricing from columns W and Z
"""
import openpyxl
import json

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'

def extract_materials(sheet):
    """Extract materials from Material sheet columns W, Z"""
    print(f"\n=== EXTRACTING MATERIALS ===")
    
    items = []
    
    # Column W = Name (23), Z = Price (26)
    for row_idx in range(2, sheet.max_row + 1):  # Skip header row 1
        name = sheet.cell(row=row_idx, column=23).value  # Column W
        area = sheet.cell(row=row_idx, column=24).value  # Column X (area m²)
        price = sheet.cell(row=row_idx, column=26).value  # Column Z
        
        # Stop at empty row
        if not name or name == '':
            break
        
        # Skip header rows
        if 'نوع' in str(name) or 'متریال' in str(name):
            continue
        
        item = {
            "code": f"MAT-{row_idx-1}",
            "description": str(name).strip(),
            "unit": "متر مربع",  # All materials are sold per square meter
            "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0,
            "category": "Panel",
            "persianNames": [str(name).strip()],
            "isActive": True
        }
        items.append(item)
        print(f"  {item['code']}: {item['description']} - {item['unitPrice']:,.0f} Rials")
    
    print(f"  Total: {len(items)}")
    return items

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    materials = extract_materials(wb['Material'])
    
    # Read existing pricing tables to keep CNC, edge banding, etc.
    with open('api/src/data/pricing-tables.json', 'r', encoding='utf-8') as f:
        existing_data = json.load(f)
    
    # Update only materials
    existing_data['materials'] = materials
    
    # Write back
    with open('api/src/data/pricing-tables.json', 'w', encoding='utf-8') as f:
        json.dump(existing_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Materials updated!")
    print(f"   Total materials: {len(materials)}")

if __name__ == "__main__":
    main()

