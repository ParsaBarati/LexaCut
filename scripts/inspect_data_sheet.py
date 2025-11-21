#!/usr/bin/env python3
"""
Inspect the Data sheet to find the pricing tables
"""
import openpyxl

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'

def inspect_data_sheet(wb):
    """Inspect the Data sheet in detail"""
    print(f"\n{'='*60}")
    print(f"SHEET: Data")
    print(f"{'='*60}")
    
    sheet = wb['Data']
    
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    print(f"\nScanning for data tables...\n")
    
    # Scan all rows to find tables
    for row_idx in range(1, min(200, sheet.max_row + 1)):
        row_data = []
        has_data = False
        for col_idx in range(1, min(15, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None and value != "":
                has_data = True
            if value is None:
                value = ""
            elif isinstance(value, str):
                value = value[:40]  # Truncate long strings
            row_data.append(f"[{col_idx}]{value}")
        
        if has_data:
            print(f"Row {row_idx}: {' | '.join(row_data)}")

def inspect_fittings_data(wb):
    """Inspect the FittingsData sheet"""
    print(f"\n{'='*60}")
    print(f"SHEET: FittingsData")
    print(f"{'='*60}")
    
    sheet = wb['FittingsData']
    
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    print(f"\nFirst 50 rows:\n")
    
    for row_idx in range(1, min(50, sheet.max_row + 1)):
        row_data = []
        has_data = False
        for col_idx in range(1, min(15, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None and value != "":
                has_data = True
            if value is None:
                value = ""
            elif isinstance(value, str):
                value = value[:40]  # Truncate long strings
            row_data.append(f"[{col_idx}]{value}")
        
        if has_data:
            print(f"Row {row_idx}: {' | '.join(row_data)}")

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    inspect_data_sheet(wb)
    inspect_fittings_data(wb)

if __name__ == "__main__":
    main()

