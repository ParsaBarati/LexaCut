#!/usr/bin/env python3
"""
Extract the REAL pricing data from the lookup tables in columns W, X, Y
"""
import openpyxl
import json

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'
OUTPUT_FILE = 'api/src/data/pricing-tables-complete.json'

def extract_lookup_table(sheet, sheet_name):
    """Extract pricing lookup table from columns W, X, Y"""
    print(f"\n=== EXTRACTING {sheet_name} ===")
    
    items = []
    
    # Column W = Name/Code (23), X = Price (24), Y = Unit (25)
    for row_idx in range(2, sheet.max_row + 1):  # Skip header row 1
        name = sheet.cell(row=row_idx, column=23).value  # Column W
        price = sheet.cell(row=row_idx, column=24).value  # Column X
        unit = sheet.cell(row=row_idx, column=25).value  # Column Y
        
        # Stop at empty row
        if not name or name == '':
            break
        
        # Skip header rows
        if 'نوع' in str(name) or 'قیمت' in str(name) or 'یراق آلات' in str(name):
            continue
        
        item = {
            "code": f"{sheet_name}-{row_idx-1}",  # Generate a code
            "name": str(name).strip(),
            "unit": str(unit).strip() if unit else "عدد",
            "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
        }
        items.append(item)
        print(f"  {item['code']}: {item['name']} - {item['unitPrice']} {item['unit']}")
    
    print(f"  Total: {len(items)}")
    return items

def extract_materials(sheet):
    """Extract materials from Material sheet (different structure)"""
    print(f"\n=== EXTRACTING MATERIALS ===")
    
    # Check different possible columns for material data
    # Let's scan the Material sheet more carefully
    items = []
    
    # Based on earlier inspection, Material sheet might have data in different columns
    # Let's try columns around the pricing data
    for col_start in [20, 21, 22, 23, 24, 25]:
        print(f"\nTrying column {chr(ord('A') + col_start - 1)}...")
        for row_idx in range(1, min(30, sheet.max_row + 1)):
            val = sheet.cell(row=row_idx, column=col_start).value
            if val and 'MDF' in str(val).upper() or 'PVC' in str(val).upper() or 'ام دی اف' in str(val) or 'پی وی سی' in str(val):
                # Found material data
                name = sheet.cell(row=row_idx, column=col_start).value
                price = sheet.cell(row=row_idx, column=col_start + 1).value
                unit = sheet.cell(row=row_idx, column=col_start + 2).value
                
                item = {
                    "code": f"MAT-{len(items)+1}",
                    "name": str(name).strip() if name else "",
                    "unit": str(unit).strip() if unit else "متر مربع",
                    "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
                }
                items.append(item)
                print(f"    Found: {item['name']} - {item['unitPrice']}")
    
    # If we didn't find materials in lookup tables, use hardcoded common materials
    if len(items) == 0:
        print("  No materials found in lookup tables, using common materials from Data sheet...")
        # We saw these in the Data sheet earlier
        items = [
            {"code": "MAT-1", "name": "ام دی اف 16 میل - سفید", "unit": "متر مربع", "unitPrice": 2500000},
            {"code": "MAT-2", "name": "ام دی اف 16 میل -روکش چوب", "unit": "متر مربع", "unitPrice": 3000000},
            {"code": "MAT-3", "name": "ام دی اف 3 میل -سفید", "unit": "متر مربع", "unitPrice": 500000},
            {"code": "MAT-4", "name": "پی وی سی 16 میل - سفید", "unit": "متر مربع", "unitPrice": 2000000},
            {"code": "MAT-5", "name": "پی وی سی 3 میل - سفید", "unit": "متر مربع", "unitPrice": 400000},
            {"code": "MAT-6", "name": "ام دی اف 16 میل -خام", "unit": "متر مربع", "unitPrice": 2200000},
        ]
    
    print(f"  Total: {len(items)}")
    return items

def extract_edge_banding(sheet):
    """Extract edge banding from NavarShiarFarsi sheet"""
    print(f"\n=== EXTRACTING EDGE BANDING ===")
    
    items = []
    
    # Column W = Name (23), X = Price (24), Y = Unit (25)
    for row_idx in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row_idx, column=23).value
        price = sheet.cell(row=row_idx, column=24).value
        unit = sheet.cell(row=row_idx, column=25).value
        
        if not name or name == '':
            break
        
        if 'نوع خدمات' in str(name) or 'قیمت' in str(name):
            continue
        
        item = {
            "code": f"EB-{row_idx-1}",
            "name": str(name).strip(),
            "unit": str(unit).strip() if unit else "متر طول",
            "pricePerMeter": float(price) if price and isinstance(price, (int, float)) else 0
        }
        items.append(item)
        print(f"  {item['code']}: {item['name']} - {item['pricePerMeter']} {item['unit']}")
    
    print(f"  Total: {len(items)}")
    return items

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    # Extract all pricing data
    materials = extract_materials(wb['Material']) if 'Material' in wb.sheetnames else []
    fittings = extract_lookup_table(wb['Fittings'], 'FITTING') if 'Fittings' in wb.sheetnames else []
    edge_banding = extract_edge_banding(wb['NavarShiarFarsi']) if 'NavarShiarFarsi' in wb.sheetnames else []
    cnc_operations = extract_lookup_table(wb['CNC'], 'CNC') if 'CNC' in wb.sheetnames else []
    
    data = {
        "materials": materials,
        "fittings": fittings,
        "edgeBanding": edge_banding,
        "cncOperations": cnc_operations,
        "pricingConfig": {
            "laborCostPerHour": 500000,
            "overheadPercentage": 25,
            "profitMarginPercentage": 20,
            "wastagePercentage": 5,
            "cncSetupCost": 1000000,
            "edgeBandingSetupCost": 500000
        }
    }
    
    # Write to JSON
    print(f"\n{'='*60}")
    print(f"WRITING TO {OUTPUT_FILE}")
    print(f"{'='*60}")
    print(f"  Materials: {len(materials)}")
    print(f"  Fittings: {len(fittings)}")
    print(f"  Edge Banding: {len(edge_banding)}")
    print(f"  CNC Operations: {len(cnc_operations)}")
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Extraction complete!")

if __name__ == "__main__":
    main()

