#!/usr/bin/env python3
"""
Extract the actual pricing lookup tables from the sheets
The VBA code shows that each sheet has a lookup table in columns around W:X
"""
import openpyxl
import json

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'
OUTPUT_FILE = 'api/src/data/pricing-tables-complete.json'

def extract_lookup_table(sheet, start_col, name):
    """Extract pricing lookup table from specific columns"""
    print(f"\n=== EXTRACTING {name} LOOKUP TABLE ===")
    print(f"  Scanning columns starting at {start_col}...")
    
    items = []
    start_col_idx = ord(start_col) - ord('A') + 1
    
    # Scan rows for data
    for row_idx in range(1, sheet.max_row + 1):
        code = sheet.cell(row=row_idx, column=start_col_idx).value
        name_val = sheet.cell(row=row_idx, column=start_col_idx + 1).value
        unit = sheet.cell(row=row_idx, column=start_col_idx + 2).value if start_col_idx + 2 <= sheet.max_column else None
        price = sheet.cell(row=row_idx, column=start_col_idx + 3).value if start_col_idx + 3 <= sheet.max_column else None
        
        # Skip headers and empty rows
        if code and name_val and isinstance(code, (str, int)) and isinstance(name_val, str):
            # Skip header rows
            if str(code).lower() in ['code', 'کد', 'ردیف']:
                continue
            if 'شرح' in str(name_val) or 'نام' in str(name_val):
                continue
                
            item = {
                "code": str(code).strip(),
                "name": str(name_val).strip(),
                "unit": str(unit).strip() if unit else "عدد",
                "unitPrice": float(price) if price and isinstance(price, (int, float)) else 0
            }
            items.append(item)
            print(f"  Row {row_idx}: {item['code']} - {item['name']} - {item['unitPrice']}")
    
    print(f"  Total extracted: {len(items)}")
    return items

def scan_all_columns(sheet, sheet_name):
    """Scan all columns to find pricing data"""
    print(f"\n=== SCANNING ALL COLUMNS IN {sheet_name} ===")
    
    # Check columns U through Z (columns 21-26)
    for col_idx in range(21, 30):
        col_letter = chr(ord('A') + col_idx - 1)
        print(f"\n  Column {col_letter}:")
        found_data = False
        for row_idx in range(1, min(30, sheet.max_row + 1)):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val:
                print(f"    Row {row_idx}: {val}")
                found_data = True
        if not found_data:
            print(f"    (empty)")

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    # Check CNC sheet first
    if 'CNC' in wb.sheetnames:
        scan_all_columns(wb['CNC'], 'CNC')
        
        # Try extracting from column W (column 23)
        cnc_operations = extract_lookup_table(wb['CNC'], 'W', 'CNC Operations')
    
    # Check Fittings sheet
    if 'Fittings' in wb.sheetnames:
        scan_all_columns(wb['Fittings'], 'Fittings')
        fittings = extract_lookup_table(wb['Fittings'], 'W', 'Fittings')
    
    # Check NavarShiarFarsi sheet
    if 'NavarShiarFarsi' in wb.sheetnames:
        scan_all_columns(wb['NavarShiarFarsi'], 'NavarShiarFarsi')
        edge_banding = extract_lookup_table(wb['NavarShiarFarsi'], 'W', 'Edge Banding')

if __name__ == "__main__":
    main()

