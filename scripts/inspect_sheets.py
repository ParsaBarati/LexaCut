#!/usr/bin/env python3
"""
Inspect the Excel sheets to understand their structure
"""
import openpyxl

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'

def inspect_sheet(wb, sheet_name, max_rows=20):
    """Inspect a specific sheet"""
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    
    sheet = wb[sheet_name]
    
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    print(f"\nFirst {max_rows} rows:\n")
    
    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(10, sheet.max_column + 1)):  # First 10 columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                value = ""
            elif isinstance(value, str):
                value = value[:30]  # Truncate long strings
            row_data.append(f"{col_idx}:{value}")
        print(f"Row {row_idx}: {' | '.join(row_data)}")

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    # Inspect the key sheets
    sheets_to_inspect = ['Material', 'Fittings', 'NavarShiarFarsi', 'CNC', 'BoreshKari']
    
    for sheet_name in sheets_to_inspect:
        if sheet_name in wb.sheetnames:
            inspect_sheet(wb, sheet_name)
        else:
            print(f"\nSheet '{sheet_name}' not found!")

if __name__ == "__main__":
    main()

