#!/usr/bin/env python3
"""
Extract pricing tables from the "All" sheet which should contain all pricing data
"""
import openpyxl
import json

FILE_NAME = 'ANALIZ-MALI-GHARARDAD-BIM.xlsm'
OUTPUT_FILE = 'api/src/data/pricing-tables-complete.json'

def extract_pricing_from_all_sheet(wb):
    """Extract all pricing data from the 'All' sheet"""
    print("\n=== EXTRACTING FROM 'All' SHEET ===")
    
    if 'All' not in wb.sheetnames:
        print("ERROR: 'All' sheet not found")
        return None
        
    sheet = wb['All']
    print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    
    # Scan the sheet to find pricing tables
    print("\nScanning for pricing data...")
    
    materials = []
    fittings = []
    edge_banding = []
    cnc_operations = []
    
    current_section = None
    
    for row_idx in range(1, min(500, sheet.max_row + 1)):
        # Read first few columns to identify sections
        col1 = sheet.cell(row=row_idx, column=1).value
        col2 = sheet.cell(row=row_idx, column=2).value
        col3 = sheet.cell(row=row_idx, column=3).value
        col4 = sheet.cell(row=row_idx, column=4).value
        col5 = sheet.cell(row=row_idx, column=5).value
        
        # Check for section headers
        if col1 and isinstance(col1, str):
            if 'متریال' in col1 or 'مواد' in col1:
                current_section = 'materials'
                print(f"Found materials section at row {row_idx}")
                continue
            elif 'یراق' in col1 or 'fittings' in col1.lower():
                current_section = 'fittings'
                print(f"Found fittings section at row {row_idx}")
                continue
            elif 'نوار' in col1:
                current_section = 'edge_banding'
                print(f"Found edge banding section at row {row_idx}")
                continue
            elif 'cnc' in col1.lower() or 'سی ان سی' in col1:
                current_section = 'cnc'
                print(f"Found CNC section at row {row_idx}")
                continue
        
        # Extract data based on current section
        if current_section and col2 and col3:  # We have code and name
            item = {
                "code": str(col2).strip() if col2 else "",
                "name": str(col3).strip() if col3 else "",
                "unit": str(col4).strip() if col4 else "",
                "unitPrice": float(col5) if col5 and isinstance(col5, (int, float)) else 0
            }
            
            if current_section == 'materials':
                materials.append(item)
            elif current_section == 'fittings':
                fittings.append(item)
            elif current_section == 'edge_banding':
                item["pricePerMeter"] = item.pop("unitPrice")
                edge_banding.append(item)
            elif current_section == 'cnc':
                cnc_operations.append(item)
    
    print(f"\nExtracted:")
    print(f"  - Materials: {len(materials)}")
    print(f"  - Fittings: {len(fittings)}")
    print(f"  - Edge Banding: {len(edge_banding)}")
    print(f"  - CNC Operations: {len(cnc_operations)}")
    
    return {
        "materials": materials,
        "fittings": fittings,
        "edgeBanding": edge_banding,
        "cncOperations": cnc_operations,
        "pricingConfig": {
            "laborCostPerHour": 0,
            "overheadPercentage": 0,
            "profitMarginPercentage": 0,
            "wastagePercentage": 5,
            "cncSetupCost": 0,
            "edgeBandingSetupCost": 0
        }
    }

def main():
    print(f"Loading {FILE_NAME}...")
    wb = openpyxl.load_workbook(FILE_NAME, data_only=True)
    
    print(f"\nAvailable sheets: {wb.sheetnames}")
    
    # First try the 'All' sheet
    data = extract_pricing_from_all_sheet(wb)
    
    if not data or (not data['materials'] and not data['fittings'] and not data['cncOperations']):
        print("\nNo data found in 'All' sheet. Let me manually list ALL rows from the CNC sheet:")
        
        # Let's manually check CNC sheet
        if 'CNC' in wb.sheetnames:
            print("\n=== CNC SHEET FULL DUMP ===")
            sheet = wb['CNC']
            for row_idx in range(1, min(100, sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, 8):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val:
                        row_data.append(f"[{col_idx}]{val}")
                if row_data:
                    print(f"Row {row_idx}: {' | '.join(row_data)}")
        
        return
    
    # Write to JSON
    print(f"\n=== WRITING TO {OUTPUT_FILE} ===")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\nExtraction complete!")

if __name__ == "__main__":
    main()

